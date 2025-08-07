import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from typing import Dict, Any, Optional
from datetime import datetime


class BaseReport:
    def __init__(self):
        self.workbook_name = '' # 文件名
        self.worksheet_name = '' # 工作表名
        self.file_path = f'{self.workbook_name}.xlsx'  # 默认文件路径
        self.color = '000000' # 字体、边框颜色
        self.name = '宋体'
        self.title = '' # 报表标题
        # 标题样式
        self.title_font = Font(name=self.name, size=18, bold=True, color=self.color)
        # 表头及小标题样式
        self.header_font = Font(name=self.name, size=11, bold=True, color=self.color)
        # 数据样式
        self.data_font = Font(name=self.name, size=11, bold=False, color=self.color)
        # 对齐
        self.center_align = Alignment(horizontal='center', vertical='center')
        # 边框
        self.border = Border(
            left=Side(style='thin', color=self.color),
            right=Side(style='thin', color=self.color),
            top=Side(style='thin', color=self.color),
            bottom=Side(style='thin', color=self.color)
        )
        self.column_num = 1
        # self.workbook = Workbook()
        # self.worksheet = self.workbook.active
        # 添加填充色
        self.title_fill = PatternFill(fill_type=None)
        self.header_fill = PatternFill(fill_type=None)

    def _create_workbook(self) -> tuple[Workbook, Worksheet]:
        workbook = Workbook()
        worksheet = workbook.active
        # 这句其实没啥用，主要是为了消除类型注解报错
        if worksheet is None:
            worksheet = workbook.create_sheet()
        return (workbook, worksheet)

    # 添加报表标题
    def _add_title(self, ws: Worksheet) -> int:
        # 设置待合并单元格边框
        # for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=self.column_num):
        #     for cell in row:
        #         cell.border = self.border
        for cell in ws[1][0:self.column_num]:
            cell.border = self.border
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=self.column_num)
        title_cell: Cell = ws['A1']
        title_cell.value = self.title
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.center_align
        return 2  # 返回下一行的位置

    # 添加小标题
    def _add_header_title(self, ws: Worksheet, header_title: str, start_row: int) -> int:
        current_row = start_row
        # for row in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=self.column_num):
        #     for cell in row:
        #         cell.border = self.border
        for cell in ws[current_row][0:self.column_num]:
            cell.border = self.border
        ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=self.column_num)
        header_cell: Cell = ws[f'A{current_row}']
        header_cell.value = header_title
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = self.center_align
        return current_row + 1

    # 添加表头尾内容，注意这里直接用dataframe批量写入了，没做任何判断
    def _add_header_footer_info(self, ws: Worksheet, header_footer_data: pd.DataFrame, start_row: int) -> int:
        current_row = start_row
        if header_footer_data is None or header_footer_data.empty:
            return current_row
        rows_cnt, cols_cnt = header_footer_data.shape
        rows = dataframe_to_rows(header_footer_data, index=False, header=False)
        for row in rows:
            ws.append(row)
        last_row = current_row + rows_cnt - 1
        for row in ws.iter_rows(min_row=current_row, max_row=last_row, min_col=1, max_col=cols_cnt):
            for cell in row:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_align
        return last_row + 1
    
    # 添加报表数据
    def _add_data(self, ws: Worksheet, data: pd.DataFrame, start_row: int) -> int:
        current_row = start_row
        if data is None or data.empty:
            return current_row
        rows_cnt, cols_cnt = data.shape
        rows = dataframe_to_rows(data, index=False, header=False)
        for row in rows:
            ws.append(row)
        last_row = current_row + rows_cnt - 1
        for row in ws.iter_rows(min_row=current_row, max_row=last_row, min_col=1, max_col=cols_cnt):
            for cell in row:
                cell.font = self.data_font
                cell.border = self.border
                cell.alignment = self.center_align
        return last_row + 1
    
    def _adjust_column_widths(self, ws: Worksheet, strategy: str = 'fixed', custom_width: int = 0) -> None:
        """
        调整工作表列宽
        
        Args:
            ws: 工作表对象
            strategy: 调整策略 ('auto', 'custom', 'fixed')
            min_width: 最小列宽
            max_width: 最大列宽  
            custom_widths: 自定义列宽字典，如 {'A': 15, 'B': 20}
        """
        try:
            # if strategy == 'custom' and custom_width:
            #     # 使用自定义列宽
            #     for column_letter, width in custom_widths.items():
            #         ws.column_dimensions[column_letter].width = width
            
            # elif strategy == 'fixed':
                # 固定列宽
            default_width = custom_width if custom_width > 0 else 20
            for i in range(1, self.column_num + 1):
                column_letter = chr(64 + i)
                ws.column_dimensions[column_letter].width = default_width
            
            # else:  # 'auto' - 自动计算
            #     for column in ws.columns:
            #         max_length = 0
            #         column_letter = column[0].column_letter
                    
            #         for cell in column:
            #             if cell.value:
            #                 # 计算内容长度，考虑中文字符
            #                 content = str(cell.value)
            #                 # 中文字符计算（占用更多空间）
            #                 chinese_chars = len([c for c in content if '\u4e00' <= c <= '\u9fff'])
            #                 length = len(content) + chinese_chars * 0.5
                            
            #                 max_length = max(max_length, length)
                    
            #         # 应用最小和最大宽度限制
            #         adjusted_width = max(min_width, min(max_length + 2, max_width))
            #         ws.column_dimensions[column_letter].width = adjusted_width
                    
        except:
            # 异常处理：设置默认列宽
            for i in range(1, self.column_num + 1):
                column_letter = chr(64 + i)
                ws.column_dimensions[column_letter].width = 20


class TQReportGenerator(BaseReport):
    def __init__(self):
        super().__init__()
    def generate_report(self, device_name: str, report_data: dict[str, list[pd.DataFrame]]) -> bytes:
        self.column_num = 6
        wb, ws = self._create_workbook()
        self.title = f"提取车间自控报表--{device_name}"
        # 报表标题
        current_row = self._add_title(ws)
        # header_info
        header_info = report_data['header'][0]
        current_row = self._add_header_footer_info(ws, header_info, current_row)
        # main_info
        main_info = report_data['main']
        header_title = ['一次参数设置', '一次煎煮记录']
        for title, data in zip(header_title, main_info):
            current_row = self._add_header_title(ws, title, current_row)
            current_row = self._add_data(ws, data, current_row)
        # footer_info
        footer_title = '其他信息'
        footer_info = report_data['footer'][0]
        current_row = self._add_header_title(ws, footer_title, current_row)
        self._add_header_footer_info(ws, footer_info, current_row)
        # 调整列宽
        self._adjust_column_widths(ws, custom_width=20)
        
        # 保存到内存
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

class ReportGeneratorFactory:
    GENERATOR_MAPPING: dict[str, type] = {
        "T_TQ_Batch_Archive": TQReportGenerator
    }

    @classmethod
    def create_report_generator(cls, table_name: str) -> type:
        generator_class = cls.GENERATOR_MAPPING.get(table_name)
        if generator_class:
            return generator_class()
        else:
            raise ValueError(f"不支持的报表类型: {table_name}")

# 工厂方法，供外部调用
def get_report(report_data: dict[str, list[pd.DataFrame]], table_name: str, device_name: str) -> bytes:
    generator = ReportGeneratorFactory.create_report_generator(table_name)
    report = generator.generate_report(device_name, report_data)
    return report

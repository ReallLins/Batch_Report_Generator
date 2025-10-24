from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import re


WORKBOOK_PATH = "Report_Template/Report_Template_1#提取罐.xlsx"
wb = load_workbook(WORKBOOK_PATH)

def get_order_list(wb: Workbook) -> list[str]:
    defined_name_list = list(wb.defined_names)
    def extract_number(name):
        match = re.search(r'block_(\d+)', name)
        return int(match.group(1)) if match else float('inf')
    return sorted(defined_name_list, key=extract_number)

def get_title_data(ws: Worksheet, area: str):
    cell = ws[area]
    return cell.value if cell.value is not None else "None"

def get_content_data(ws: Worksheet, area: str) -> list[str]:
    data = []
    rows = ws[area]
    for row in rows:
        data_row = []
        for cell in row:
            data_row.append(cell.value if cell.value is not None else "None")
        data.append(data_row)
    return data

# defined_list = get_order_list(wb)
# print(defined_list)

# for name in defined_list:
#     print(f"=== {name} ===")
#     defined = wb.defined_names[name]
#     for title, coord in defined.destinations:
#         print(f"{title}: {coord}")
#         sheet = wb[title]
#         cell = sheet[coord]
#         print(cell.value)

